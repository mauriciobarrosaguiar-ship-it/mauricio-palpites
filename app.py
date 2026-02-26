# -*- coding: utf-8 -*-
"""
Mauricio Palpites (Mobile Web) - Streamlit (mobile-first)
- UI em cards (estilo app)
- Busca jogos na internet (football-data.org + TheSportsDB fallback)
- Top 3 do dia, filtro por % e multi-ligas
- Exporta CSV (PT-BR) e Excel (.xlsx) com Ranking + Cores por confiança

OBS: token do football-data pode ser sobrescrito via:
- Streamlit Secrets: FOOTBALL_DATA_TOKEN
- Variável de ambiente: FOOTBALL_DATA_TOKEN
"""

import os
import io
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional

import requests
import pandas as pd
import numpy as np
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# Config
# =========================
APP_TITLE = "⚽ Mauricio Palpites"
OUT_CSV_NAME = "palpites_saida.csv"

FD_BASE = "https://api.football-data.org/v4"
TSDB_BASE = "https://www.thesportsdb.com/api/v1/json/3"  # chave pública "3" (limitada)

# Token default (pode sobrescrever em Secrets/Env)
DEFAULT_FD_TOKEN = "2f1cc1a3dcde4cfd8ef56002a5284302"

# Heurística (não é garantia)
K_MANDANTE = 0.08
K_EMPATE_BASE = 0.26


# =========================
# Logs
# =========================
def log(msg: str):
    st.session_state.setdefault("logs", [])
    ts = datetime.now().strftime("%H:%M:%S")
    st.session_state["logs"].append(f"[{ts}] {msg}")


def clear_logs():
    st.session_state["logs"] = []


# =========================
# Token / headers
# =========================
def fd_token() -> Optional[str]:
    # 1) Secrets
    try:
        t = st.secrets.get("FOOTBALL_DATA_TOKEN", None)
        if t:
            return t
    except Exception:
        pass
    # 2) Env
    t = os.getenv("FOOTBALL_DATA_TOKEN")
    if t:
        return t
    # 3) Default (pedido pelo usuário)
    return DEFAULT_FD_TOKEN


def fd_headers() -> Dict[str, str]:
    tok = fd_token()
    if not tok:
        return {}
    return {"X-Auth-Token": tok}


def fd_enabled() -> bool:
    return bool(fd_headers().get("X-Auth-Token"))


# =========================
# Catálogo
# =========================
@st.cache_data(ttl=60 * 60, show_spinner=False)
def load_catalog() -> List[Dict]:
    items: List[Dict] = []

    # football-data (com token)
    if fd_enabled():
        try:
            r = requests.get(f"{FD_BASE}/competitions", headers=fd_headers(), timeout=20)
            r.raise_for_status()
            data = r.json()
            for c in data.get("competitions", []):
                name = c.get("name") or ""
                cid = c.get("id")
                code = c.get("code") or ""
                if not cid or not name:
                    continue
                items.append({"provider": "FD", "key": f"FD:{cid}", "name": name, "code": code, "id": cid})
            log("Catálogo football-data carregado.")
        except Exception as e:
            log(f"AVISO: falha ao carregar catálogo football-data: {e}")

    # TSDB fallback
    try:
        r = requests.get(f"{TSDB_BASE}/all_leagues.php", timeout=20)
        r.raise_for_status()
        data = r.json()
        for l in data.get("leagues", []):
            if (l.get("strSport") or "").lower() != "soccer":
                continue
            name = l.get("strLeague") or ""
            lid = l.get("idLeague")
            if not lid or not name:
                continue
            items.append({"provider": "TSDB", "key": f"TSDB:{lid}", "name": name, "id": lid})
        log("Catálogo TheSportsDB carregado.")
    except Exception as e:
        log(f"ERRO: não consegui carregar catálogo TheSportsDB: {e}")

    # Dedup
    seen = set()
    out = []
    for it in items:
        k = (it["provider"], it["name"].strip().lower())
        if k in seen:
            continue
        seen.add(k)
        out.append(it)

    return sorted(out, key=lambda x: (x["provider"], x["name"]))


def search_catalog(items: List[Dict], q: str) -> List[Dict]:
    q = (q or "").strip().lower()
    if not q:
        return items
    terms = [t for t in q.replace("-", " ").split() if t]
    if not terms:
        return items

    def score(it: Dict) -> int:
        name = it["name"].lower()
        sc = 0
        for t in terms:
            if t in name:
                sc += 2
            if name.startswith(t):
                sc += 1
        return sc

    ranked = sorted(items, key=lambda it: score(it), reverse=True)
    return [it for it in ranked if score(it) > 0]


# =========================
# Jogos
# =========================
def fd_matches(competition_id: int, date_from: str, date_to: str) -> List[Dict]:
    url = f"{FD_BASE}/competitions/{competition_id}/matches"
    params = {"dateFrom": date_from, "dateTo": date_to}
    r = requests.get(url, headers=fd_headers(), params=params, timeout=25)
    r.raise_for_status()
    return r.json().get("matches", [])


def tsdb_events_next(league_id: str) -> List[Dict]:
    url = f"{TSDB_BASE}/eventsnextleague.php"
    r = requests.get(url, params={"id": league_id}, timeout=25)
    r.raise_for_status()
    return (r.json() or {}).get("events", []) or []


# =========================
# Probabilidade (heurística)
# =========================
def clamp(x: float, a=0.01, b=0.99) -> float:
    return max(a, min(b, x))


def simple_confidence(seed: int) -> float:
    rng = np.random.default_rng(seed)
    base = 55 + rng.normal(0, 12)
    base += K_MANDANTE * 100
    return float(np.clip(base, 5, 95))


def best_bet_from_conf(conf: float) -> Tuple[str, float]:
    p_home = clamp(0.33 + (conf - 50) / 130)
    p_draw = clamp(K_EMPATE_BASE - (conf - 50) / 260, 0.08, 0.32)
    p_away = clamp(1 - p_home - p_draw, 0.12, 0.62)

    s = p_home + p_draw + p_away
    p_home, p_draw, p_away = p_home / s, p_draw / s, p_away / s

    probs = {"casa vence": p_home, "empate": p_draw, "fora vence": p_away}
    best = max(probs.items(), key=lambda kv: kv[1])
    return best[0], float(best[1])


def palpite_humano(best_bet: str, mandante: str, visitante: str) -> str:
    b = (best_bet or "").strip().lower()
    if b == "casa vence":
        return f"{mandante} ganha"
    if b == "fora vence":
        return f"{visitante} ganha"
    if b == "empate":
        return "Empate (X)"
    return best_bet


def to_portuguese(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = pd.DataFrame({
        "data": df["date"],
        "fonte": df["provider"].map({"FD": "football-data", "TSDB": "TheSportsDB"}).fillna(df["provider"]),
        "competicao": df["league"],
        "mandante": df["home"],
        "visitante": df["away"],
        "palpite": [palpite_humano(b, h, a) for b, h, a in zip(df["best_bet"], df["home"], df["away"])],
        "probabilidade_pct": (df["p_best_bet"].astype(float) * 100.0).round(1),
        "confianca_0a100": pd.to_numeric(df["confidence"], errors="coerce").round(1),
        "placar_mais_provavel": df["most_likely_score"],
        "origem": df.get("data_source", ""),
    })
    return out


# =========================
# Excel (bytes): ranking + cores
# =========================
def excel_bytes_formatado(df_pt: pd.DataFrame) -> bytes:
    if df_pt is None or df_pt.empty:
        raise ValueError("Sem dados para exportar.")

    df2 = df_pt.copy()
    df2["confianca_0a100"] = pd.to_numeric(df2["confianca_0a100"], errors="coerce")
    df2["probabilidade_pct"] = pd.to_numeric(df2["probabilidade_pct"], errors="coerce")
    df2 = df2.sort_values(["confianca_0a100", "probabilidade_pct"], ascending=False).reset_index(drop=True)
    df2.insert(0, "Ranking", range(1, len(df2) + 1))

    headers = list(df2.columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mauricio Palpites"

    bg_dark = "0B3D2E"
    header_green = "198754"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tcell = ws.cell(row=1, column=1)
    tcell.value = "Mauricio Palpites"
    tcell.font = Font(size=18, bold=True, color="FFFFFF")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    tcell.fill = PatternFill(start_color=bg_dark, end_color=bg_dark, fill_type="solid")
    ws.row_dimensions[1].height = 34

    header_fill = PatternFill(start_color=header_green, end_color=header_green, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    ws.row_dimensions[2].height = 24

    for r_idx, row in enumerate(df2.itertuples(index=False), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions

    fill_high = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")
    fill_mid  = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")
    fill_low  = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")

    conf_col_idx = headers.index("confianca_0a100") + 1 if "confianca_0a100" in headers else None
    if conf_col_idx:
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=conf_col_idx).value
            try:
                v = float(v)
            except Exception:
                continue
            if v >= 75:
                fill = fill_high
            elif v >= 60:
                fill = fill_mid
            else:
                fill = fill_low
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, min(ws.max_row, 500) + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# Geração
# =========================
def build_rows_for_item(item: Dict, days_ahead: int) -> pd.DataFrame:
    today = datetime.utcnow().date()
    date_from = today.strftime("%Y-%m-%d")
    date_to = (today + timedelta(days=days_ahead)).strftime("%Y-%m-%d")

    provider = item["provider"]
    league_name = item["name"]
    rows = []

    if provider == "FD":
        cid = int(item["id"])
        try:
            matches = fd_matches(cid, date_from, date_to)
            if not matches:
                log(f"[{league_name}] sem jogos na janela ({days_ahead} dias).")
                return pd.DataFrame()

            for m in matches:
                status = (m.get("status") or "").upper()
                if status not in ("SCHEDULED", "TIMED"):
                    continue
                utc = m.get("utcDate") or ""
                date = utc[:10] if utc else ""
                home = (m.get("homeTeam") or {}).get("name", "")
                away = (m.get("awayTeam") or {}).get("name", "")
                seed = abs(hash((league_name, date, home, away))) % (2**32)
                conf = simple_confidence(seed)
                best, pbest = best_bet_from_conf(conf)

                rows.append({
                    "date": date,
                    "provider": "FD",
                    "league": league_name,
                    "home": home,
                    "away": away,
                    "best_bet": best,
                    "p_best_bet": pbest,
                    "confidence": conf,
                    "most_likely_score": "1-0" if best == "casa vence" else ("0-1" if best == "fora vence" else "1-1"),
                    "data_source": "football-data.org",
                })
        except Exception as e:
            log(f"ERRO football-data ({league_name}): {e}")
            return pd.DataFrame()

    else:
        lid = str(item["id"])
        try:
            evs = tsdb_events_next(lid)
            if not evs:
                log(f"[{league_name}] sem eventos retornados (TSDB).")
                return pd.DataFrame()

            for ev in evs:
                date = ev.get("dateEvent") or ""
                try:
                    d = datetime.strptime(date, "%Y-%m-%d").date()
                    if d < today or d > today + timedelta(days=days_ahead):
                        continue
                except Exception:
                    pass

                home = ev.get("strHomeTeam") or ""
                away = ev.get("strAwayTeam") or ""
                seed = abs(hash((league_name, date, home, away))) % (2**32)
                conf = simple_confidence(seed)
                best, pbest = best_bet_from_conf(conf)

                rows.append({
                    "date": date,
                    "provider": "TSDB",
                    "league": league_name,
                    "home": home,
                    "away": away,
                    "best_bet": best,
                    "p_best_bet": pbest,
                    "confidence": conf,
                    "most_likely_score": "1-0" if best == "casa vence" else ("0-1" if best == "fora vence" else "1-1"),
                    "data_source": "TheSportsDB",
                })
        except Exception as e:
            log(f"ERRO TSDB ({league_name}): {e}")
            return pd.DataFrame()

    return pd.DataFrame(rows)


def generate(selected_items: List[Dict], days_ahead: int, min_prob: float) -> pd.DataFrame:
    if not selected_items:
        return pd.DataFrame()

    dfs = [build_rows_for_item(it, days_ahead) for it in selected_items]
    out = pd.concat([d for d in dfs if d is not None and not d.empty], ignore_index=True) if dfs else pd.DataFrame()
    if out.empty:
        return out

    out = out.sort_values(["confidence", "p_best_bet"], ascending=False).reset_index(drop=True)
    out = out[out["p_best_bet"] * 100.0 >= min_prob].reset_index(drop=True)
    return out


# =========================
# UI (Mobile-first)
# =========================
st.set_page_config(page_title="Mauricio Palpites", page_icon="⚽", layout="centered")

st.markdown(
    """
    <style>
      .block-container { padding-top: 0.8rem; padding-bottom: 2.0rem; }
      .stButton>button, .stDownloadButton>button { width: 100%; border-radius: 14px; padding: 0.85rem 1rem; font-weight: 800; }
      .chip { display:inline-block; padding:6px 10px; border-radius:999px; font-weight:700; font-size:0.85rem; }
      .chip-green{ background:#dff0d8; }
      .chip-yellow{ background:#fcf8e3; }
      .chip-red{ background:#f2dede; }
      .card {
        background: white;
        border-radius: 18px;
        padding: 14px 14px;
        box-shadow: 0 8px 20px rgba(0,0,0,0.06);
        border: 1px solid rgba(0,0,0,0.06);
        margin-bottom: 12px;
      }
      .t1 { font-size: 1.05rem; font-weight: 900; margin: 0; }
      .t2 { font-size: 0.92rem; opacity: 0.85; margin: 4px 0 0 0; }
      .t3 { font-size: 0.95rem; margin: 10px 0 0 0; font-weight: 800; }
      .muted { opacity: 0.72; font-size: 0.85rem; }
      .section-title { margin-top: 0.6rem; margin-bottom: 0.2rem; font-weight: 900; }
    </style>
    """,
    unsafe_allow_html=True
)

st.title(APP_TITLE)
st.caption("Abra no celular e use **Adicionar à tela inicial** para virar app.")

# Config
with st.expander("⚙️ Configuração", expanded=True):
    q = st.text_input("Buscar Competição/Liga", value="")
    catalog = load_catalog()
    filtered = search_catalog(catalog, q)

    options = [f"[{it['provider']}] {it['name']}" for it in filtered]
    key_to_item = {f"[{it['provider']}] {it['name']}": it for it in filtered}

    selected_opt = st.multiselect("Selecione 1 ou mais ligas", options=options[:250], default=[])
    selected_items = [key_to_item[o] for o in selected_opt if o in key_to_item]

    c1, c2 = st.columns(2)
    with c1:
        days_ahead = st.slider("Dias à frente", 1, 30, 10)
    with c2:
        min_prob = st.slider("Só acima de (%)", 0, 95, 70)

    run = st.button("🎯 Gerar palpites")

# Run
if run:
    clear_logs()
    log(f"Gerando | ligas={len(selected_items)} | dias={days_ahead} | filtro>={min_prob}%")
    with st.spinner("Buscando jogos e calculando..."):
        raw = generate(selected_items, days_ahead, min_prob)

    if raw.empty:
        st.warning("Sem jogos na janela (ou nenhum palpite passou no filtro). Aumente os dias ou diminua o %.")
    else:
        df_pt = to_portuguese(raw)
        st.session_state["df_pt"] = df_pt
        st.success(f"Encontrado: {len(df_pt)} palpites")

df_pt = st.session_state.get("df_pt", pd.DataFrame())

def conf_chip(conf: float) -> str:
    try:
        c = float(conf)
    except Exception:
        c = 0.0
    if c >= 75:
        return '<span class="chip chip-green">Confiança Alta</span>'
    if c >= 60:
        return '<span class="chip chip-yellow">Confiança Média</span>'
    return '<span class="chip chip-red">Confiança Baixa</span>'

def money_prob(prob: float) -> str:
    try:
        p = float(prob)
    except Exception:
        p = 0.0
    return f"{p:.1f}%"

# Results as cards
if df_pt is not None and not df_pt.empty:
    st.markdown("### ✅ Resultados")
    df_sorted = df_pt.sort_values(["confianca_0a100", "probabilidade_pct"], ascending=False).reset_index(drop=True)

    # Top 3
    st.markdown("### 🏆 Top 3 apostas do dia")
    top3 = df_sorted.head(3)
    for i, r in top3.iterrows():
        st.markdown(
            f"""
            <div class="card">
              <p class="t1">{i+1}. {r['mandante']} x {r['visitante']}</p>
              <p class="t2">{r['competicao']} • {r['data']}</p>
              <p class="t3">Palpite: {r['palpite']}</p>
              <p class="muted">Probabilidade: <b>{money_prob(r['probabilidade_pct'])}</b> • Confiança: <b>{r['confianca_0a100']}</b> {conf_chip(r['confianca_0a100'])}</p>
            </div>
            """,
            unsafe_allow_html=True
        )

    st.markdown("### 📋 Lista completa")
    for _, r in df_sorted.iterrows():
        st.markdown(
            f"""
            <div class="card">
              <p class="t1">{r['mandante']} x {r['visitante']}</p>
              <p class="t2">{r['competicao']} • {r['data']} • {r['fonte']}</p>
              <p class="t3">{r['palpite']}</p>
              <p class="muted">Probabilidade: <b>{money_prob(r['probabilidade_pct'])}</b> • Confiança: <b>{r['confianca_0a100']}</b> {conf_chip(r['confianca_0a100'])}</p>
            </div>
            """,
            unsafe_allow_html=True
        )

    st.markdown("### ⬇️ Exportar")
    csv_bytes = df_sorted.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig").encode("utf-8-sig")
    st.download_button("📄 Baixar CSV", data=csv_bytes, file_name=OUT_CSV_NAME, mime="text/csv")

    try:
        xlsx_bytes = excel_bytes_formatado(df_sorted)
        st.download_button(
            "📊 Baixar Excel (.xlsx)",
            data=xlsx_bytes,
            file_name="mauricio_palpites.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption("O Excel inclui **Ranking** e **cores por nível de confiança**.")
    except Exception as e:
        st.error(f"Falha ao gerar Excel: {e}")

# Logs
with st.expander("📄 Logs / Detalhes", expanded=False):
    if st.button("🧹 Limpar logs"):
        clear_logs()
    logs = st.session_state.get("logs", [])
    if logs:
        st.code("\n".join(logs), language="text")
    else:
        st.caption("Sem logs no momento.")

st.caption("Fontes: football-data.org (com token) e TheSportsDB (fallback).")
